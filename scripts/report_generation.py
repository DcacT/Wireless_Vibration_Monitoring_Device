from openpyxl import Workbook, load_workbook
from numpy.fft import fft as fft
import numpy as np
import os
import csv
import matplotlib.pyplot as plt
amp_array_x = []
amp_array_y = []
amp_array_z = []
freq_array = []
def append_to_excel(file_name):
    xls_file_name = "./data/reports/"+file_name+".xlsx"
    """
    Append data to an Excel file. If the file doesn't exist, create it with the required headers.

    :param file_name: Name of the Excel file.
    :param data: List of data to append as a new row.
    """
    if os.path.exists(xls_file_name):
        os.remove(xls_file_name)
        print(f"{xls_file_name} has been deleted.")
    else:
        print(f"{xls_file_name} does not exist.")
    headers = [
        "Order", "Time", 
        "Vib_Data_x","Vib_Data_y","Vib_Data_z","",
        "FFT_x","FFT_y","FFT_z","",
        "Freq_x","Freq_y","Freq_z","",
        "Amp_x", "Amp_y", "Amp_z","","",

        "Samples", "Frequency Resolution"
    ]

    if not os.path.exists(xls_file_name):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Vibration Data"

        # Add headers
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)
        # Load the existing workbook
        csv_file_name = './data/raw/'+file_name+'.csv'
        with open(csv_file_name, mode='r') as csvfile:
            csvreader = csv.reader(csvfile)
            counter = 0
            # Append rows from the CSV to the Excel sheet
            start_time = 0
            end_time = 0
            fft_x_data = []
            fft_y_data = []
            fft_z_data = []
            freq = 1000
            freq_half = freq/2
            freq_half_counter = 0
            freq_resolution = (counter + 1 )/freq


            for row in csvreader:
                if counter == 0:
                    start_time = float(row[1])
                else:
                    end_time = float(row[1])
                print(row)
                sheet.append((float(row[0]),float(row[1]),float(row[2])/330,float(row[3])/330,float(row[4])/330))  # Append each row from CSV to Excel        
                fft_x_data.append(float(row[2]))
                fft_y_data.append(float(row[3]))
                fft_z_data.append(float(row[4]))
                counter +=1
            
            
            fft_x_array = fft(fft_x_data)
            fft_y_array = fft(fft_y_data)
            fft_z_array = fft(fft_z_data)



            for i in range(len(fft_x_array)):
                #fft
                sheet.cell(row = 2+i, column = 7).value = str(fft_x_array[i])
                sheet.cell(row = 2+i, column = 8).value = str(fft_y_array[i])
                sheet.cell(row = 2+i, column = 9).value = str(fft_z_array[i])

                #freq
                freq_array.append(i/freq_resolution)
                sheet.cell(row = 2+i, column = 11).value = str(i/freq_resolution)
                sheet.cell(row = 2+i, column = 12).value = str(i/freq_resolution)
                sheet.cell(row = 2+i, column = 13).value = str(i/freq_resolution)
                
                #amp
                amp_array_x.append(np.abs((fft_x_array[i]) /(0.5 *counter + 1) / 2))
                amp_array_y.append(np.abs((fft_y_array[i]) /(0.5 *counter + 1) / 2))
                amp_array_z.append(np.abs((fft_z_array[i]) /(0.5 *counter + 1) / 2))

                sheet.cell(row = 2+i, column = 15).value = str(np.abs((fft_x_array[i]) /(0.5 *counter + 1) / 2))
                sheet.cell(row = 2+i, column = 16).value = str(np.abs((fft_y_array[i]) /(0.5 *counter + 1) / 2))
                sheet.cell(row = 2+i, column = 17).value = str(np.abs((fft_z_array[i]) /(0.5 *counter + 1) / 2))

            
            total_time = end_time - start_time
            # freq_resolution = (counter + 1 )/((counter + 1 )/total_time)
            freq_resolution = (counter + 1 )/freq

            sheet.cell(row = 2, column = 20).value = counter+1
            sheet.cell(row = 2, column = 21).value = freq_resolution

            

        # Save the updated workbook
        workbook.save(xls_file_name)
        print(f"Data appended to {xls_file_name} successfully.")


# Example usage:
file_name = "Data_2024-11-15_23-19-20"
# Example data to append
append_to_excel(file_name)




# Create a figure with 3 subplots for x, y, and z
fig, axs = plt.subplots(3, 1, figsize=(10, 6), sharex=True)

# Plot for the x component
axs[0].plot(freq_array[1:len(amp_array_x)//2], amp_array_x[1:len(amp_array_x)//2], label="FFT Amplitude (x)", color='blue')
axs[0].set_ylabel("Amplitude")
axs[0].set_title("FFT of the Signal (X)")
axs[0].grid(True)
axs[0].legend()

# Plot for the y component
axs[1].plot(freq_array[1:len(amp_array_x)//2], amp_array_y[1:len(amp_array_y)//2], label="FFT Amplitude (y)", color='green')
axs[1].set_ylabel("Amplitude")
axs[1].set_title("FFT of the Signal (Y)")
axs[1].grid(True)
axs[1].legend()

# Plot for the z component
axs[2].plot(freq_array[1:len(amp_array_x)//2], amp_array_z[1:len(amp_array_z)//2], label="FFT Amplitude (z)", color='red')
axs[2].set_xlabel("Frequency (Hz)")
axs[2].set_ylabel("Amplitude")
axs[2].set_title("FFT of the Signal (Z)")
axs[2].grid(True)
axs[2].legend()

# Show the plot
plt.tight_layout()  # To avoid overlap of labels
plt.show()